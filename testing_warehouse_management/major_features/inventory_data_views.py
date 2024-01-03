import openpyxl
import re
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.shortcuts import render
from openpyxl.styles import Border, Font, Alignment, Side
from io import BytesIO
from django.contrib import messages
from . reports_views import validating_period_id
from . models import *

def inventory_data(request, accounting_period_id):
    # Default redirecting to 'inventory_data' view
    # with current accounting_period_id
    if accounting_period_id == 0:
        accounting_period_id = AccoutingPeriod.objects.select_related('warehouse_management_method').latest('id').pk
        return HttpResponseRedirect(reverse('inventory_data', kwargs={'accounting_period_id': accounting_period_id}))

    if request.method == "POST":
        chosen_period_id_param = request.POST.get("accounting_period", None)
        accounting_period_id = validating_period_id(chosen_period_id_param)

        # Redirecting to 'inventory_data' view
        # with the chosen period id in POST request
        return HttpResponseRedirect(reverse('inventory_data', kwargs={'accounting_period_id': accounting_period_id}))

    # For rendering select input in the form
    accounting_periods = AccoutingPeriod.objects.select_related('warehouse_management_method').all()

    # For displaying each accounting period in the table
    accounting_periods_inventory = AccountingPeriodInventory.objects.select_related('accounting_period_id', 'product_id').filter(
        accounting_period_id = accounting_period_id
    )

    # For summarizing factors
    periods_summarizing_factors = accounting_periods_inventory.aggregate(
        total_starting_quantity = Sum("starting_quantity"),
        total_starting_inventory = Sum("starting_inventory"),
        total_import_quantity = Sum("import_quantity"),
        total_import_inventory = Sum("import_inventory"),
        total_export_quantity = Sum("total_quantity_export"),
        total_products_cogs = Sum("total_cogs"),
        total_ending_quantity = Sum("ending_quantity"),
        total_ending_inventory = Sum("ending_inventory"),
        total_products_revenue = Sum("total_revenue"),
        total_gross_profits = Sum("total_revenue") - Sum("total_cogs")
    )

    context = {
        'accounting_periods': accounting_periods,
        'chosen_period_id': accounting_period_id,
        'accounting_periods_inventory': accounting_periods_inventory,

        'total_starting_quantity': periods_summarizing_factors['total_starting_quantity'],
        'total_starting_inventory': periods_summarizing_factors['total_starting_inventory'],
        'total_import_quantity': periods_summarizing_factors['total_import_quantity'],
        'total_import_inventory': periods_summarizing_factors['total_import_inventory'],
        'total_export_quantity': periods_summarizing_factors['total_export_quantity'],
        'total_products_cogs': periods_summarizing_factors['total_products_cogs'],
        'total_ending_quantity': periods_summarizing_factors['total_ending_quantity'],
        'total_ending_inventory': periods_summarizing_factors['total_ending_inventory'],
        'total_products_revenue': periods_summarizing_factors['total_products_revenue'],
        'total_gross_profits': periods_summarizing_factors['total_gross_profits']
    }

    if accounting_periods_inventory.count() == 0:
        non_record_msg = "Không có dữ liệu"
        context["non_record_msg"] = non_record_msg

    return render(request, "major_features/inventory_data/inventory_data.html", context)

def export_data_to_excel(request, accounting_period_id):
    filename = filename_handling(request)
    if filename is None:
        # Add fail message
        messages.add_message(request, messages.ERROR, "Tên file không hợp lệ")
        return HttpResponseRedirect(reverse('inventory_data'))

    # Get all products period inventory
    accounting_periods_inventory = AccountingPeriodInventory.objects.select_related('accounting_period_id', 'product_id').filter(
        accounting_period_id = accounting_period_id
    )
    # Summarizing factors
    periods_summarizing_factors = accounting_periods_inventory.aggregate(
        total_starting_quantity = Sum("starting_quantity"),
        total_starting_inventory = Sum("starting_inventory"),
        total_import_quantity = Sum("import_quantity"),
        total_import_inventory = Sum("import_inventory"),
        total_export_quantity = Sum("total_quantity_export"),
        total_products_export_value = Sum("total_cogs"),
        total_ending_quantity = Sum("ending_quantity"),
        total_ending_inventory = Sum("ending_inventory"),
        total_products_revenue = Sum("total_revenue"),
        total_products_cogs = Sum("total_cogs"),
        total_gross_profits = Sum("total_revenue") - Sum("total_cogs")
    )

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "HTK kỳ kế toán hiện tại"

    # Populating data
    # Populating header
    worksheet = populating_header(worksheet)

    # Populating body
    worksheet = populating_body(worksheet, accounting_periods_inventory)

    # Populating footer
    worksheet = populating_footer(worksheet, periods_summarizing_factors)

    # Create a temporary memory object using BytesIO
    excelfile = BytesIO()

    # Save the new created one workbook to the excelfile
    workbook.save(excelfile)

    return download_excelfile(request, excelfile, filename)

def populating_header(worksheet):
    headers = [
        "STT",
        "Sản phẩm",
        "Danh mục",
        "Tồn kho đầu kỳ",
        "Nhập kho trong kỳ",
        "Xuất kho trong kỳ"
        "Tồn kho cuối kỳ",
        "Doanh thu",
        "GVHB",
        "Lợi nhuận"
    ]
    thin = Side(border_style="thin", color="000000")
    # STT cell header
    cell_A1 = worksheet["A1"]
    cell_A1.value = "STT"
    cell_A1.alignment = Alignment(horizontal="center", vertical="center")
    cell_A1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_A1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Product cell header
    cell_B1 = worksheet["B1"]
    cell_B1.value = "Sản phẩm"
    cell_B1.alignment = Alignment(vertical="center")
    cell_B1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_B1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Category cell header
    cell_C1 = worksheet["C1"]
    cell_C1.value = "Danh mục"
    cell_C1.alignment = Alignment(vertical="center")
    cell_C1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_C1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Starting Inventory cell header
    cell_D1 = worksheet["D1"]
    cell_D1.value = "Tồn kho đầu kỳ"
    cell_D1.alignment = Alignment(horizontal="center")
    cell_D1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_D1.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Import Inventory cell header
    cell_F1 = worksheet["F1"]
    cell_F1.value = "Nhập kho trong kỳ"
    cell_F1.alignment = Alignment(horizontal="center")
    cell_F1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_F1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Export cell header
    cell_H1 = worksheet["H1"]
    cell_H1.value = "Xuất kho trong kỳ"
    cell_H1.alignment = Alignment(horizontal="center")
    cell_H1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_H1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Ending inventory cell header
    cell_J1 = worksheet["J1"]
    cell_J1.value = "Tồn kho cuối kỳ"
    cell_J1.alignment = Alignment(horizontal="center")
    cell_J1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_J1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Revenue cell header
    cell_L1 = worksheet["L1"]
    cell_L1.value = "Doanh thu"
    cell_L1.alignment = Alignment(vertical="center")
    cell_L1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_L1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # COGS cell header
    cell_M1 = worksheet["M1"]
    cell_M1.value = "GVHB"
    cell_M1.alignment = Alignment(vertical="center")
    cell_M1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_M1.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Gross Profits header
    cell_N1 = worksheet["N1"]
    cell_N1.value = "Lợi nhuận"
    cell_N1.alignment = Alignment(vertical="center")
    cell_N1.font = Font(name="Times New Roman", bold=True, size=12)
    cell_N1.border = Border(top=thin, right=thin, left=thin, bottom=thin)

    # Merging rows
    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:B2")
    worksheet.merge_cells("C1:C2")
    worksheet.merge_cells("L1:L2")
    worksheet.merge_cells("M1:M2")
    worksheet.merge_cells("N1:N2")

    # Merging columns
    worksheet.merge_cells("D1:E1")
    worksheet.merge_cells("F1:G1")
    worksheet.merge_cells("H1:I1")
    worksheet.merge_cells("J1:K1")

    cell_D2 = worksheet["D2"]
    cell_D2.value = "Số lượng"
    cell_D2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_D2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_E2 = worksheet["E2"]
    cell_E2.value = "Giá trị"    
    cell_E2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_E2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    cell_F2 = worksheet["F2"]
    cell_F2.value = "Số lượng"
    cell_F2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_F2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_G2 = worksheet["G2"]
    cell_G2.value = "Giá trị"
    cell_G2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_G2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    cell_H2 = worksheet["H2"]
    cell_H2.value = "Số lượng"
    cell_H2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_H2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_I2 = worksheet["I2"]
    cell_I2.value = "Giá trị"
    cell_I2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_I2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    cell_J2 = worksheet["J2"]
    cell_J2.value = "Số lượng"
    cell_J2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_J2.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_K2 = worksheet["K2"]
    cell_K2.value = "Giá trị"
    cell_K2.font = Font(name="Times New Roman", bold=True, size=12)
    cell_K2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    return worksheet

def populating_body(worksheet, accounting_periods):
    thin = Side(border_style="thin", color="000000")
    row_num = 3
    counter = 1
    for _, period in enumerate(accounting_periods, 1):
        row = [
            counter,
            period.product_id.name,
            period.product_id.category_name.name,
            period.starting_quantity,
            period.starting_inventory,
            period.import_quantity,
            period.import_inventory,
            period.total_quantity_export,
            period.total_cogs,
            period.ending_quantity,
            period.ending_inventory,
            period.total_revenue,
            period.total_cogs,
            period.total_revenue - period.total_cogs
        ]
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(name="Times New Roman", size=12)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Bolding & centering the STT cells
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Times New Roman", size=12, bold=True)
            
            # Bolding the products name cells
            if col_num == 2:
                cell.font = Font(name="Times New Roman", size=12, bold=True)

            # If encountering cell numbers
            if col_num >= 3:
                cell.number_format = '#,##0'
        
        row_num += 1
        counter += 1
    
    return worksheet

def populating_footer(worksheet, summarize_factors):
    thin = Side(border_style="thin", color="000000")
    START_FACTOR_VALUE_INDEX = 4

    # Col span the summarize cell
    max_row = worksheet.max_row
    footer_row = max_row + 1
    footer_summarize_cell = worksheet.cell(row=footer_row, column=1)
    footer_summarize_cell.value = "Tổng"
    footer_summarize_cell.alignment = Alignment(horizontal="right")
    footer_summarize_cell.font = Font(name="Times New Roman", bold=True, size=12)
    footer_summarize_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    col_span_group = f"A{footer_row}:C{footer_row}"
    worksheet.merge_cells(col_span_group)

    col_num = START_FACTOR_VALUE_INDEX

    for factor, value in summarize_factors.items():
        cell = worksheet.cell(row=footer_row, column=col_num)
        cell.value = value
        cell.font = Font(name="Times New Roman")
        cell.number_format = "#,##0"
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        col_num += 1

    return worksheet

def filename_handling(request):
    # Default filename
    filename = "HTK"
    # Get filename param
    filename_param = request.GET.get("filename", None)
    if filename_param:
        filename = filename_param

    filename += ".xlsx"
    # Validating filename
    if re.match(r"^[a-zA-Z0-9_\- ]+\.xlsx$", filename) is None:
        return None
    
    return filename

def download_excelfile(request, excelfile, filename):
    # Create the HttpResponse object with the appropriate headers
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f"attachment; filename={filename}"

    # Write the workbook data as a bytes-like object
    # that was created before using BytesIO
    response.write(excelfile.getvalue())
    
    return response
